'''
Created on Aug 17, 2019

@author: ye.wen
'''

import pandas as pd
import numpy as np
import math
import urllib.request  as urllib2 
import openpyxl as pl
from openpyxl.styles import colors
import Tools.data_cleanup_helpers as data_cleanup_helpers

#declaration of variables
anatomy = None
model = None
model_data_only = None

def update_cell(sheet, cell, val, color = colors.RED):
    sheet[cell] = val
    sheet[cell].font = pl.styles.Font(color=color)



def __update_model_from_one_table__(source_table, source_col, dest_col):
    
    return __update_sheet_from_one_table__(model, model_data_only, anatomy, source_table, source_col, dest_col)

'''
def __update_model_from_one_table__(source_table, source_col, dest_col):    
    for item in anatomy.index:
        for i in source_table.index:
            if not(type(source_table.loc[i,0]) is str): continue
            if item in source_table.loc[i,0]:
                val = source_table.loc[i,source_col]
                cell =  dest_col + str( anatomy.loc[item, 'Row Number'])
                original_val = model_data_only[cell].value
                if type(original_val) is float:
                    original_val = round(model_data_only[cell].value,1)
                yield [item, val, original_val]
                update_cell(model, cell, val, color = colors.BLUE)
                break
'''

    


def __update_sheet_from_one_table__(sheet, sheet_data_only, row_ref, source_table, source_col, dest_col, source_col_history = None, dest_col_history = None, updating_if_not_matching = False):    
    for item in row_ref.index:
        for i in source_table.index:
            if not(type(source_table.loc[i,0]) is str): continue



            if item in source_table.loc[i,0]:

                to_update = True
                
                if source_col_history != None:
                    source_val_history = source_table.loc[i,source_col_history]
                    cell_history =  dest_col_history + str( row_ref.loc[item, 'Row Number'])
                    dest_val_history = sheet_data_only[cell_history].value
                    if source_val_history != dest_val_history:
                        t = 'No match: ' + str(source_val_history) + ' vs. ' + str(dest_val_history)
                        to_update = updating_if_not_matching
                    else:
                        t = 'Match'

                val = source_table.loc[i,source_col]
                cell =  dest_col + str( row_ref.loc[item, 'Row Number'])
                original_val = sheet_data_only[cell].value
                
                
                if type(original_val) is float:
                    original_val = round(sheet_data_only[cell].value,1)
                
                if source_col_history != None:
                    yield [item, val, original_val, t]
                else:
                    yield [item, val, original_val]
                
                
                if to_update: update_cell(sheet, cell, val, color = colors.BLUE)
                break




def __update_model_from_one_matrix__(source_matrix, dest_col):    
    for item in anatomy.index:
        pos = item.find('__')
        if pos<0: continue
        
        index = item[0:pos]
        col_name = item[pos+2:]
        
        source_matrix.reset_index(drop=True, inplace = True)
        source_matrix.columns = range(source_matrix.shape[1])
        
        for i in source_matrix.index:
            if not(type(source_matrix.loc[i,0]) is str): continue
                
            if index in source_matrix.loc[i,0]:
                for j in source_matrix.columns:
                    if not(type(source_matrix.loc[0,j]) is str): continue
                    if col_name in source_matrix.loc[0,j]:
                        
                        val = source_matrix.loc[i,j]
                        cell =  dest_col + str( anatomy.loc[item, 'Row Number']) 
                        original_val = model_data_only[cell].value

                        if type(original_val) is float:
                            original_val = round(model_data_only[cell].value,1)
                        yield( [item, val, original_val])
                        update_cell(model, cell, val, color = colors.BLUE)
                        break

def __update_model_from_tables_and_matrixes__(source_tables, source_table_cols, source_matrixes, dest_col):
    for idx, table in enumerate(source_tables):
        a = list(__update_model_from_one_table__(table, source_table_cols[idx], dest_col))
        yield pd.DataFrame(a)

    for matrix in source_matrixes:
        a = list(__update_model_from_one_matrix__(matrix, dest_col))
        yield pd.DataFrame(a)



'''
Update specific rows in a specific excel sheet, with one table. Ideal for updating BS or CF  
'''
def update_sheet(workbook_filename, sheet_name, row_start, row_end, source_table, source_col, dest_col,  source_col_history = None, dest_col_history = None, updating_if_not_matching = False, output_filename = 'output.xlsx'):
    wb_data_only = pl.load_workbook(workbook_filename, data_only=True)
    wb = pl.load_workbook(workbook_filename)
    sheet = wb[sheet_name]
    sheet_data_only = wb_data_only[sheet_name]
    
    row_ref = []
    
    i = row_start
    while i<=row_end:
        cell =  'A'+ str(i)
        item = sheet_data_only[cell].value
        if type(item) is str:
            if item != '':
                row_ref.append([item, i])
        i += 1
    
    row_ref = pd.DataFrame(row_ref).rename(columns = {0: 'Item', 1:'Row Number'}).set_index('Item')
    
    df = __update_sheet_from_one_table__(sheet, sheet_data_only, row_ref, source_table, source_col, dest_col, source_col_history = source_col_history,  dest_col_history = dest_col_history, updating_if_not_matching = updating_if_not_matching)

    df = list(df)
    df = pd.DataFrame(df)
    if source_col_history == None:
        df = df.rename(columns = {0:'Item', 1:'New Value', 2:'Old Value'}).set_index('Item')
    else:
        df = df.rename(columns = {0:'Item', 1:'New Value', 2:'Old Value', 3:'Historical Value'}).set_index('Item')
    
    df['Diff'] =  (df['New Value']/df['Old Value']).map(lambda x: '{:,.1%}'.format(x))
    
    
    wb.save(output_filename)
    return df


'''
Update excel sheet 'Model' with items defined in sheet 'Anatomy'
'''
def update_model(model_filename, source_tables, source_table_cols, matrixes, dest_col, output_filename = 'output.xlsx'):

    global anatomy
    global model
    global model_data_only
    
    wb_data_only = pl.load_workbook(model_filename, data_only=True)
    ws = wb_data_only['Anatomy']
    anatomy = pd.DataFrame(ws.values)
    new_header = anatomy.iloc[0] 
    anatomy = anatomy[1:] 
    anatomy.columns = new_header
    anatomy.set_index('Item', inplace= True)
    anatomy = anatomy.dropna(axis=1, how='all')
    anatomy = anatomy.dropna(axis=0, how='all')
    
    wb = pl.load_workbook(model_filename)
    model = wb['Model']
    model_data_only = wb_data_only['Model']
    
    
    dfs = __update_model_from_tables_and_matrixes__(source_tables, source_table_cols, matrixes, dest_col)
    dfs = list(dfs)
    df = pd.concat(dfs, axis=0)
    if not df.empty:
        df = df.rename(columns = {0:'Item', 1:'New Value', 2:'Old Value'}).set_index('Item')
        df['Diff'] =  (df['New Value']/df['Old Value']-1).map(lambda x: '{:,.1%}'.format(x))

    
    wb.save(output_filename)
    return df

if __name__ == '__main__':
    pass